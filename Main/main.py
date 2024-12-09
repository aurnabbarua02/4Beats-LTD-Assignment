from openpyxl import load_workbook
from selenium import webdriver
import time
import datetime

file = load_workbook('Excel.xlsx')
sheetName = datetime.datetime.today().strftime('%A')
print('Today is', sheetName,'. So, sheet name \'', sheetName,'\' will be modified.')
sheet = file[sheetName]

driver=webdriver.Edge()
driver.maximize_window()
driver.get("https://www.google.com/")

i=1
for row in sheet.iter_rows():
    keywordRow=0
    keywordCol=0
    for col in row:
        searchKey = driver.find_element('name', "q")
        searchKey.clear()
        keyword = f'Keyword{i}'

        if col.value == keyword:
            keywordRow=col.row
            keywordCol=col.column
            i+=1
            searchKey.send_keys(sheet.cell(row=keywordRow, column=keywordCol+1).value)
            time.sleep(3)
            suggestions = driver.find_elements('css selector', 'li.sbct')
            longest_suggestion = max(suggestions, key=lambda suggestion: len(suggestion.text))
            sheet.cell(row=keywordRow, column=keywordCol+2).value = longest_suggestion.text
            sheet.cell(row=keywordRow, column=keywordCol+3).value = sheet.cell(row=keywordRow, column=keywordCol+1).value
            break

file.save('Excel.xlsx')
print('Completed')